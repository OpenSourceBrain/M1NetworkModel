"""
utils.py 

Useful functions related to the parameters file, eg. create params file from excel table 

Contributors: salvadordura@gmail.com
"""
from neuron import h

def getSecName(sec):
	fullSecName = sec.name().split('.')[1]
	if '[' in fullSecName:  # if section is array element
		secNameTemp = fullSecName.split('[')[0]
		secIndex = int(fullSecName.split('[')[1].split(']')[0])
		secName = secNameTemp+'_'+str(secIndex) #(secNameTemp,secIndex)
	else:
		secName = fullSecName
		secIndex = -1
	return secName

def importCellFromTemplate(fileName, cellName):
	''' Import cell from HOC template into framework format (dict of sections, with geom, topol, mechs, syns)'''
	h.load_file(fileName)
	cell = getattr(h, cellName)(0,19,0)  # arguments correspond to zloc, type and id -- remove in future (not used internally)

	print 'Loading cell from template: '+fileName
	secDic = {}
	for sec in cell.allsec(): 
		# create new section dict with name of section
		secName = getSecName(sec)
		secDic[secName] = {'geom': {}, 'topol': {}, 'mechs': {}, 'syns': {}}  # create dictionary to store sec info

		# store geometry properties
		standardGeomParams = ['L', 'nseg', 'diam', 'Ra', 'cm']
		secDir = dir(sec)
		for geomParam in standardGeomParams:
			if geomParam in secDir:
				secDic[secName]['geom'][geomParam] = sec.__getattribute__(geomParam)

		# store 3d geometry
		numPoints = int(h.n3d())
		if numPoints: points = []
		for ipoint in range(numPoints):
			x = h.x3d(ipoint)
			y = h.y3d(ipoint)
			z = h.z3d(ipoint)
			diam = h.diam3d(ipoint)
			points.append((x, y, z, diam))
		secDic[secName]['geom']['pt3d'] = points

		# store mechanisms
		ignoreMechs = ['dist']
		mechDic = {}
		for mech in dir(sec(0.5)):  
			if h.ismembrane(mech) and mech not in ignoreMechs:  # check if membrane mechanism
				mechDic[mech] = {}  # create dic for mechanism properties
				props = [prop.replace('_'+mech, '') for prop in dir(sec(0.5).__getattribute__(mech)) if prop.endswith('_'+mech)]
				propVals = []
				for prop in props:
					propVals = [seg.__getattribute__(mech).__getattribute__(prop) for seg in sec]
					if len(set(propVals)) == 1:
						propVals = propVals[0] 
					mechDic[mech][prop] = propVals
		secDic[secName]['mechs'] = mechDic

		# add synapses 
		# for now read fixed params, but need to find way to read only synapse params
		syns = {}
		synParams = ['e', 'tau1', 'tau2']
		for seg in sec:
			for isyn,syn in enumerate(seg.point_processes()):
				synName = 'syn_'+ str(isyn)
				syns[synName] = {}
				syns[synName]['type'] = syn.hname().split('[')[0]
				syns[synName]['loc'] = seg.x
				for synParam in synParams:
					try:
						syns[synName][synParam] = syn.__getattribute__(synParam)
					except:
						pass
		secDic[secName]['syns'] = syns

		# store topology (keep at the end since h.SectionRef messes remaining loop)
		secRef = h.SectionRef(sec=sec)
		if secRef.has_parent():
			secDic[secName]['topol']['parentSec'] = getSecName(secRef.parent().sec)
			secDic[secName]['topol']['parentX'] = h.parent_connection()
			secDic[secName]['topol']['childX'] = h.section_orientation()
	return secDic
	
	'''# soma = {'geom': {}, 'topol': {}, 'mechs': {}, 'syns': {}, 'Izhi2007Type': 'IB'}  #  soma
	# soma['geom'] = {'diam': 14.0, 'L': 18.0}  # From Suter,2013 Fig1D
	# soma['syns']['AMPA'] = {'type': 'MyExp2SynBB', 'loc': 0.5, 'tau1': 0.05, 'tau2':5.3, 'e': 0}  # AMPA
	# soma['syns']['NMDA'] = {'type': 'MyExp2SynNMDABB', 'loc': 0.5, 'tau1NMDA': 15, 'tau2NMDA': 150, 'r': 1, 'e': 0}  # NMDA
	# soma['syns']['GABAA'] = {'type': 'MyExp2SynBB', 'loc': 0.5, 'tau1': 0.07, 'tau2': 9.1, 'e': -80}  # GABAA (fast)
	# soma['syns']['GABAB'] = {'type': 'MyExp2SynBB', 'loc': 0.5, 'tau1': 0.2, 'tau2': 20, 'e': -80}  # GABAB (slow)
	# cellProp['sections'] = {'soma': soma}  # add sections to dict
	
	dend = {'geom': {}, 'topol': {}, 'mechs': {}, 'syns': {}}  # dend properties
	dend['geom'] = {'diam': 5.0, 'L': 150.0, 'Ra': 150.0, 'cm': 1, 'pt3d': []}
	dend['topol'] = {'parentSec': 'soma', 'parentX': 1.0, 'childX': 0}
	dend['mechs']['pas'] = {'g': 0.0000357, 'e': -70} 
	dend['syns']['NMDA'] = {'type': 'Exp2Syn', 'loc': 1.0, 'tau1': 0.1, 'tau2': 1, 'e': 0}
	'''


def importConnFromExcel(fileName, sheetName):
	''' Import connectivity rules from Excel sheet'''
	import openpyxl as xl

	# set columns
	colPreTags = 0 # 'A'
	colPostTags = 1 # 'B'
	colConnFunc = 2 # 'C'
	colSyn = 3 # 'D'
	colProb = 5 # 'F'
	colWeight = 6 # 'G'
	colAnnot = 8 # 'I' 

	outFileName = fileName[:-5]+'_'+sheetName+'.py' # set output file name

	connText = """## Generated using importConnFromExcel() function in params/utils.py \n\nnetParams['connParams'] = [] \n\n"""
	
	# open excel file and sheet
	wb = xl.load_workbook(fileName)
	sheet = wb.get_sheet_by_name(sheetName)
	numRows = sheet.get_highest_row()

	with open(outFileName, 'w') as f:
		f.write(connText)  # write starting text
		for row in range(1,numRows+1):
			if sheet.cell(row=row, column=colProb).value:  # if not empty row
				print 'Creating conn rule for row ' + str(row)
				# read row values
				pre = sheet.cell(row=row, column=colPreTags).value
				post = sheet.cell(row=row, column=colPostTags).value
				func = sheet.cell(row=row, column=colConnFunc).value
				syn = sheet.cell(row=row, column=colSyn).value
				prob = sheet.cell(row=row, column=colProb).value
				weight = sheet.cell(row=row, column=colWeight).value

				# write preTags
				line = "netParams['connParams'].append({'preTags': {"
				for i,cond in enumerate(pre.split(';')):  # split into different conditions
					if i>0: line = line + ", "
					cond2 = cond.split('=')  # split into key and value
					line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
				line = line + "}" # end of preTags		

				# write postTags
				line = line + ",\n'postTags': {"
				for i,cond in enumerate(post.split(';')):  # split into different conditions
					if i>0: line = line + ", "
					cond2 = cond.split('=')  # split into key and value
					line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
				line = line + "}" # end of postTags			
				line = line + ",\n'connFunc': '" + func + "'"  # write connFunc
				line = line + ",\n'synReceptor': '" + syn + "'"  # write synReceptor
				line = line + ",\n'probability': " + str(prob)  # write prob
				line = line + ",\n'weight': " + str(weight)  # write prob
				line = line + "})"  # add closing brackets
				line = line + '\n\n' # new line after each conn rule
				f.write(line)  # write to file
				
